from __future__ import annotations

import os
from dataclasses import dataclass

import openpyxl


# Canonical column names used internally
@dataclass
class StructuredRow:
    product_name: str | None
    region: str | None
    parameter: str | None
    numeric_value: float | None
    unit: str | None
    limit_type: str | None
    notes: str | None


# Maps each client's Excel header → canonical field name.
# Add new client mappings here without touching any other file.
_COLUMN_MAPS: list[dict[str, str]] = [
    # Aurora Paints layout
    {
        "product_name": "product_name",
        "region": "region",
        "parameter": "parameter",
        "value": "numeric_value",
        "unit": "unit",
        "limit_type": "limit_type",
        "notes": "notes",
    },
    # Horizon Coatings layout
    {
        "product_line": "product_name",
        "market": "region",
        "metric": "parameter",
        "metric_value": "numeric_value",
        "metric_unit": "unit",
        "classification": "limit_type",
        "remarks": "notes",
    },
]


class ExcelParser:
    """
    Reads an .xlsx file and returns a list of normalized StructuredRow objects.
    Automatically detects the column layout by matching against known mappings.
    """

    def parse(self, file_path: str) -> list[StructuredRow]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        raw_headers = [
            (cell.value or "").strip().lower()
            for cell in next(ws.iter_rows(max_row=1))
        ]
        mapping = self._resolve_mapping(raw_headers)

        rows: list[StructuredRow] = []
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {raw_headers[i]: v for i, v in enumerate(raw_row)}
            rows.append(self._map_row(row_dict, mapping))

        wb.close()
        return rows

    # ── internal ─────────────────────────────────────────────────────────────

    def _resolve_mapping(self, headers: list[str]) -> dict[str, str]:
        """Return the column map whose keys best match the file's headers."""
        best_map, best_score = _COLUMN_MAPS[0], 0
        header_set = set(headers)
        for col_map in _COLUMN_MAPS:
            score = len(header_set & col_map.keys())
            if score > best_score:
                best_map, best_score = col_map, score
        return best_map

    def _map_row(self, row: dict, mapping: dict[str, str]) -> StructuredRow:
        canonical: dict = {v: None for v in StructuredRow.__dataclass_fields__}
        for src_col, canon_field in mapping.items():
            raw = row.get(src_col)
            if raw is not None and str(raw).strip() != "":
                canonical[canon_field] = raw
        # coerce numeric_value
        raw_num = canonical.get("numeric_value")
        if raw_num is not None:
            try:
                canonical["numeric_value"] = float(raw_num)
            except (ValueError, TypeError):
                canonical["numeric_value"] = None
        return StructuredRow(**canonical)
